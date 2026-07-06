#!/usr/bin/env python3
"""Cassiopeia — one-time converter: existing workspace files -> public/seed-data.json.

Reads (one level above the repo):
  Coffee Bags.xlsx   sheet "Bags"       -> bags   (col C font colour = legend colour;
                                          bold/RGB-coloured rows = active, rest finished)
  Coffee Bags.xlsx   sheet "Brewers"    -> brewers (merged with seed/brewers.json ids)
  Brew Ideas.xlsx    sheet "Brew Ideas" -> ideas  (canonical Recipe shape)
  My Coffee Brews.csv                   -> brews + self ratings
Plus seed/*.json (brewers, grinders, recipes) copied through.

Deterministic ids so re-running the script (and re-importing) is stable.
Requires: openpyxl.  Usage: python3 scripts/convert_seed.py
"""
import csv
import json
import re
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
WORKSPACE = REPO.parent
OUT = REPO / "public" / "seed-data.json"

SELF = {"id": "self", "name": "Me", "color": "#A85A72", "isSelf": True}


def iso(v):
    return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else (str(v).strip() if v else None)


def s(v):
    """Cell -> trimmed string or None."""
    if v is None:
        return None
    out = str(v).strip()
    return out or None


def num(v):
    if v is None:
        return None
    m = re.search(r"\d+(?:\.\d+)?", str(v))
    return float(m.group()) if m else None


def font_color(cell):
    """RGB font colour as #RRGGBB, or None for default/indexed colours."""
    c = cell.font.color
    if c and c.type == "rgb" and c.rgb and len(c.rgb) == 8:
        rgb = c.rgb[2:]
        if rgb.upper() not in ("000000", "1A1A1A"):  # default text colours, not a legend
            return f"#{rgb}"
    return None


def clean(d):
    return {k: v for k, v in d.items() if v is not None}


# --- bags + brewers from Coffee Bags.xlsx -----------------------------------

def convert_bags():
    wb = openpyxl.load_workbook(WORKSPACE / "Coffee Bags.xlsx")
    ws = wb["Bags"]
    bags = []
    for row in ws.iter_rows(min_row=2):
        (sr, roaster, name, roast_date, processing, varietal, notes, origin,
         origin_country, altitude, season, typ, sca, selection, roast,
         roaster_loc, roaster_country, frozen_serves, freeze_date, photo) = [c.value for c in row[:20]]
        if not name:
            continue
        active = bool(row[2].font.b)  # bold coffee name = current bag on the stand
        frozen = frozen_serves is not None or freeze_date is not None
        bags.append(clean({
            "id": f"bag-{int(sr) if sr else len(bags) + 1}",
            "sr": int(sr) if sr else None,
            "roaster": s(roaster) or "Unknown",
            "coffeeName": s(name),
            "roastDate": iso(roast_date),
            "processing": s(processing),
            "varietal": s(varietal),
            "notes": s(notes),
            "origin": s(origin),
            "originCountry": s(origin_country),
            "altitude": s(altitude),
            "season": s(season),
            "type": s(typ),
            "scaCupScore": num(sca),
            "selection": s(selection),
            "roast": s(roast),
            "roasterLocation": s(roaster_loc),
            "roasterCountry": s(roaster_country),
            "photo": s(photo),
            "color": font_color(row[2]),
            "finished": not active,
            "frozen": frozen,
            "frozenAmount": f"{int(frozen_serves)} serves" if isinstance(frozen_serves, (int, float)) else s(frozen_serves),
            "freezeDate": iso(freeze_date),
        }))

    seed_brewers = json.loads((REPO / "seed" / "brewers.json").read_text())
    by_name = {b["name"].lower(): b for b in seed_brewers}
    brewers = list(seed_brewers)
    for row in wb["Brewers"].iter_rows(min_row=2):
        name = s(row[1].value)
        if name and name.lower() not in by_name:
            b = {"id": re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-"), "name": name}
            brewers.append(b)
            by_name[name.lower()] = b
    return bags, brewers


# --- ideas from Brew Ideas.xlsx ----------------------------------------------

CONFIDENCE = [("certain", "high"), ("likely", "medium")]


def confidence(v):
    v = (v or "").lower()
    for key, level in CONFIDENCE:
        if v.startswith(key):
            return level
    return "low" if v else None


def split_dose_ratio(v):
    """'20g : 360g (1:18)' -> dose '20g', ratio '1:18' (falls back to raw string)."""
    v = s(v)
    if not v:
        return None, None
    dose = None
    m = re.match(r"\s*(\d+(?:\.\d+)?g)", v)
    if m:
        dose = m.group(1)
    r = re.search(r"\(?\s*(1:\d+(?:\.\d+)?)\s*\)?", v)
    ratio = r.group(1) if r else v
    return dose, ratio


def convert_ideas():
    ws = openpyxl.load_workbook(WORKSPACE / "Brew Ideas.xlsx")["Brew Ideas"]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6), start=1):
        if s(row[0].value) == "Recipe / Method":
            header_row = i
            break
    if header_row is None:
        sys.exit("Brew Ideas.xlsx: header row not found")
    ideas = []
    for row in ws.iter_rows(min_row=header_row + 1):
        (name, best_for, brewer, dose_ratio, grind, temp, steps, target,
         why, source, conf, tried, result) = [c.value for c in row[:13]]
        if not s(name):
            continue
        dose, ratio = split_dose_ratio(dose_ratio)
        steps_txt = s(steps)
        if s(why):  # BrewIdea has no `why` field — keep the insight inside steps
            steps_txt = f"{steps_txt or ''}\nWhy: {s(why)}".strip()
        ideas.append(clean({
            "id": f"idea-{len(ideas) + 1}",
            "name": s(name),
            "bestFor": s(best_for),
            "brewer": s(brewer),
            "dose": dose,
            "ratio": ratio,
            "grind": s(grind),
            "temp": s(temp),
            "steps": steps_txt,
            "targetTime": s(target),
            "source": s(source),
            "sourceConfidence": confidence(s(conf)),
            "author": None,
            "tried": bool(s(tried) and s(tried).lower().startswith("yes")),
            "result": s(result),
            "color": font_color(row[0]),
        }))
    return ideas


# --- brews + ratings from My Coffee Brews.csv --------------------------------

AXES = {  # CSV header fragment -> axis key (mouthfeel absent from the CSV)
    "Sweetness": "sweetness", "Flavour": "flavour", "Balance": "balance",
    "Aftertaste": "aftertaste", "Body": "body", "Acidity": "acidity",
    "Bitterness": "bitterness", "Fragrance": "fragrance",
}


def match_brewer(method, brewers):
    m = (method or "").lower()
    if "switch" in m:
        name = "v60 switch"
    elif "flat" in m:
        name = "flatground"
    elif "v60" in m:
        name = "v60"
    else:
        name = m
    for b in brewers:
        if b["name"].lower() == name or b["id"] == name:
            return b["id"]
    b = {"id": re.sub(r"[^a-z0-9]+", "-", m).strip("-") or "unknown", "name": method}
    brewers.append(b)
    return b["id"]


def convert_brews(bags, brewers, grinders):
    brews, ratings = [], []
    with open(WORKSPACE / "My Coffee Brews.csv", newline="", encoding="utf-8-sig") as f:
        for i, row in enumerate(csv.DictReader(f), start=1):
            coffee = (row.get("Coffee") or "").strip()
            bag = next((b for b in bags if f"{b['roaster']} {b['coffeeName']}".lower() == coffee.lower()
                        or b["coffeeName"].lower() == coffee.lower()), None)
            if bag is None:
                print(f"  ! brew row {i}: no bag matches {coffee!r} — skipped", file=sys.stderr)
                continue
            grinder = next((g["id"] for g in grinders if g["name"].lower() == (row.get("Grinder") or "").strip().lower()), None)
            brew_id = f"brew-{i}"
            brews.append(clean({
                "id": brew_id,
                "date": s(row.get("Date")),
                "bagId": bag["id"],
                "brewerId": match_brewer(row.get("Method"), brewers),
                "grinderId": grinder,
                "filter": s(row.get("Filter")),
                "roastDate": s(row.get("Roast_date")),
                "doseG": num(row.get("Dose_in_g")),
                "waterG": num(row.get("Water_out_g")),
                "tempC": num(row.get("Temp_C")),
                "totalTime": s(row.get("Total_brew_time")),
                "grind": s(row.get("Grind")),
                "pourTechnique": s(row.get("Pour_technique")),
                "notes": s(row.get("Notes")),
                "learnings": s(row.get("Learnings")),
                "withFriends": False,
                "friendIds": [],
            }))
            scores = {}
            for header, key in AXES.items():
                col = next((c for c in row if c and c.startswith(header)), None)
                v = num(row.get(col)) if col else None
                if v is not None:
                    scores[key] = v
            if scores:
                ratings.append({"id": f"rating-{brew_id}-self", "brewId": brew_id,
                                "personId": SELF["id"], "scores": scores})
    return brews, ratings


def main():
    bags, brewers = convert_bags()
    grinders = json.loads((REPO / "seed" / "grinders.json").read_text())
    recipes = json.loads((REPO / "seed" / "recipes.json").read_text())
    ideas = convert_ideas()
    brews, ratings = convert_brews(bags, brewers, grinders)
    data = {"bags": bags, "brews": brews, "ratings": ratings, "ideas": ideas,
            "recipes": recipes, "brewers": brewers, "grinders": grinders, "people": [SELF]}
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(data, indent=1, ensure_ascii=False))
    print(f"wrote {OUT.relative_to(REPO)}: " + ", ".join(f"{len(v)} {k}" for k, v in data.items()))


if __name__ == "__main__":
    main()
